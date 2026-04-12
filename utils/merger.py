import os
import uuid
import PyPDF2
from PyPDF2 import PdfWriter, PdfReader
from docx import Document
from docx.shared import Inches
import openpyxl
from PIL import Image

class FileMerger:
    """Merge multiple files of the same type"""
    
    def merge_pdfs(self, input_files, output_folder):
        """Merge multiple PDF files"""
        output_filename = f"{uuid.uuid4()}_merged.pdf"
        output_path = os.path.join(output_folder, output_filename)
        
        merger = PdfWriter()
        
        for pdf_file in input_files:
            with open(pdf_file, 'rb') as f:
                reader = PdfReader(f)
                for page in reader.pages:
                    merger.add_page(page)
        
        with open(output_path, 'wb') as output_file:
            merger.write(output_file)
        
        return output_path
    
    def merge_word_docs(self, input_files, output_folder):
        """Merge multiple Word documents"""
        output_filename = f"{uuid.uuid4()}_merged.docx"
        output_path = os.path.join(output_folder, output_filename)
        
        merged_doc = Document()
        
        for i, docx_file in enumerate(input_files):
            doc = Document(docx_file)
            
            # Add page break between documents (except for first)
            if i > 0:
                merged_doc.add_page_break()
            
            # Copy all elements
            for element in doc.element.body:
                merged_doc.element.body.append(element)
        
        merged_doc.save(output_path)
        return output_path
    
    def merge_excel_files(self, input_files, output_folder, mode='sheets'):
        """Merge multiple Excel files
        
        Args:
            input_files: List of Excel file paths
            output_folder: Output directory
            mode: 'sheets' (each file as separate sheet) or 
                  'rows' (combine all into one sheet)
        """
        output_filename = f"{uuid.uuid4()}_merged.xlsx"
        output_path = os.path.join(output_folder, output_filename)
        
        if mode == 'sheets':
            # Each file becomes a separate sheet
            wb_merged = openpyxl.Workbook()
            wb_merged.remove(wb_merged.active)  # Remove default sheet
            
            for i, excel_file in enumerate(input_files):
                wb = openpyxl.load_workbook(excel_file)
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Create new sheet with unique name
                    new_sheet_name = f"File{i+1}_{sheet_name}"
                    ws_merged = wb_merged.create_sheet(new_sheet_name)
                    
                    # Copy data
                    for row in ws.iter_rows():
                        for cell in row:
                            ws_merged[cell.coordinate].value = cell.value
            
            wb_merged.save(output_path)
        
        else:  # mode == 'rows'
            # Combine all data into one sheet
            wb_merged = openpyxl.Workbook()
            ws_merged = wb_merged.active
            ws_merged.title = "Merged Data"
            
            current_row = 1
            
            for i, excel_file in enumerate(input_files):
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                for row in ws.iter_rows(values_only=True):
                    # Skip header rows except for first file
                    if i > 0 and current_row == 1:
                        continue
                    
                    ws_merged.append(row)
                    current_row += 1
            
            wb_merged.save(output_path)
        
        return output_path
    
    def merge_images(self, input_files, output_folder, direction='vertical', spacing=0):
        """Merge multiple images
        
        Args:
            input_files: List of image file paths
            output_folder: Output directory
            direction: 'vertical' or 'horizontal'
            spacing: Pixels between images
        """
        output_filename = f"{uuid.uuid4()}_merged.png"
        output_path = os.path.join(output_folder, output_filename)
        
        images = [Image.open(f) for f in input_files]
        
        if direction == 'vertical':
            # Calculate total height and max width
            total_height = sum(img.height for img in images) + spacing * (len(images) - 1)
            max_width = max(img.width for img in images)
            
            # Create new image
            merged = Image.new('RGB', (max_width, total_height), (255, 255, 255))
            
            # Paste images
            y_offset = 0
            for img in images:
                # Convert to RGB if necessary
                if img.mode in ('RGBA', 'LA', 'P'):
                    background = Image.new('RGB', img.size, (255, 255, 255))
                    if img.mode == 'P':
                        img = img.convert('RGBA')
                    background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                    img = background
                
                merged.paste(img, (0, y_offset))
                y_offset += img.height + spacing
        
        else:  # horizontal
            # Calculate total width and max height
            total_width = sum(img.width for img in images) + spacing * (len(images) - 1)
            max_height = max(img.height for img in images)
            
            # Create new image
            merged = Image.new('RGB', (total_width, max_height), (255, 255, 255))
            
            # Paste images
            x_offset = 0
            for img in images:
                # Convert to RGB if necessary
                if img.mode in ('RGBA', 'LA', 'P'):
                    background = Image.new('RGB', img.size, (255, 255, 255))
                    if img.mode == 'P':
                        img = img.convert('RGBA')
                    background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                    img = background
                
                merged.paste(img, (x_offset, 0))
                x_offset += img.width + spacing
        
        merged.save(output_path, quality=95)
        return output_path
    
    def merge_images_grid(self, input_files, output_folder, columns=2, spacing=10):
        """Merge images in a grid layout"""
        output_filename = f"{uuid.uuid4()}_grid.png"
        output_path = os.path.join(output_folder, output_filename)
        
        images = [Image.open(f) for f in input_files]
        
        # Calculate grid dimensions
        rows = (len(images) + columns - 1) // columns
        
        # Find max dimensions for uniform sizing
        max_width = max(img.width for img in images)
        max_height = max(img.height for img in images)
        
        # Calculate total canvas size
        canvas_width = columns * max_width + (columns - 1) * spacing
        canvas_height = rows * max_height + (rows - 1) * spacing
        
        # Create canvas
        grid = Image.new('RGB', (canvas_width, canvas_height), (255, 255, 255))
        
        # Place images
        for idx, img in enumerate(images):
            row = idx // columns
            col = idx % columns
            
            x = col * (max_width + spacing)
            y = row * (max_height + spacing)
            
            # Convert to RGB if necessary
            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                img = background
            
            # Center image in cell
            x_offset = (max_width - img.width) // 2
            y_offset = (max_height - img.height) // 2
            
            grid.paste(img, (x + x_offset, y + y_offset))
        
        grid.save(output_path, quality=95)
        return output_path
    
    def merge_text_files(self, input_files, output_folder, separator="\n\n---\n\n"):
        """Merge multiple text files"""
        output_filename = f"{uuid.uuid4()}_merged.txt"
        output_path = os.path.join(output_folder, output_filename)
        
        with open(output_path, 'w', encoding='utf-8') as outfile:
            for i, txt_file in enumerate(input_files):
                if i > 0:
                    outfile.write(separator)
                
                with open(txt_file, 'r', encoding='utf-8') as infile:
                    outfile.write(infile.read())
        
        return output_path